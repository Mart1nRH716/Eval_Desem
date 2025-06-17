import os
import re
from calendar import month_name
import pandas as pd
import pyodbc
from openpyxl import load_workbook

# -------------------------------
# CONFIGURACIÓN
# -------------------------------
ARCHIVO = "diciembre_2024.xlsx"
CARPETA_REPORTES = "Reportes/UMAE"  # Ruta de los archivos
CONEXION_SQL = (
    "DRIVER={SQL Server};"
    "SERVER=localhost;"          
    "DATABASE=DBTRABAJODAS;"  # Las pruebas se realizaron en local para despues migrar las tabalas al 103    
    "UID=;" # ------------------Poner usuario  y contraseña              
    "PWD=;"           
)
# Función para obtener el mes y el año desde el nombre del archivo
def obtener_mes_anio(nombre_archivo):
    nombre_archivo = os.path.basename(nombre_archivo).lower().replace(".xlsx", "")
    partes = nombre_archivo.split("_")
    #mapeo de los meses
    meses_es = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
        "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10,
        "noviembre": 11, "diciembre": 12
    }
    mes = meses_es.get(partes[0], None)
    try:
        anio = int(partes[1])
    except (IndexError, ValueError):
        anio = None
    return mes, anio

def leer_datos(ruta_archivo):
    #Leemos el work_book, aquí no utilice pd debido a que queria extraer más elementos como el color de las celdas, además de leer los datos de acuerdo a las columnas especificadas
    wb = load_workbook(ruta_archivo, data_only=True)
    ws = wb.active
    mes, anio = obtener_mes_anio(ruta_archivo)
    registros = []

    # encabezados de indicadores están en fila 14 (index 13)
    indicadores = [ws.cell(row=14, column=col).value for col in range(45, 62)]  # AS (45) a BI (60)

    for fila in range(15, ws.max_row + 1):
        cve_pre = ws.cell(row=fila, column=43).value  # AQ
        nom_umae = ws.cell(row=fila, column=44).value  # AR
        if not cve_pre or not nom_umae:
            continue

        for idx, col in enumerate(range(45, 62)):  # columnas AS a BI
            celda = ws.cell(row=fila, column=col)
            valor = celda.value
            valor = valor if isinstance(valor, (int, float)) else -1 # El menos uno lo manejo para los datos que tienen la leyenda N/A (No aplica) ya que los datos None no es compatible en la tabla.
            indicador = indicadores[idx] if idx < len(indicadores) else f"ind_{col}" #En caso de que nos pasemos pongo otro nombre
            registros.append({
                'anio': anio,
                'mes': mes,
                'cve_pre': str(cve_pre),
                'nom_umae': str(nom_umae),
                'indicador': str(indicador).strip(),
                'valor': float(valor)
            })

    ## Total
    col_total = 62  # BJ
    indicador_total = 'total'

    for fila in range(16, 48):
        cve_pre = ws.cell(row=fila, column=43).value  # AQ
        nom_umae = ws.cell(row=fila, column=44).value  # AR
        if not cve_pre or not nom_umae:
            continue

        celda = ws.cell(row=fila, column=col_total)
        valor = celda.value
        valor = valor if isinstance(valor, (int, float)) else 0

        registros.append({
            'anio': anio,
            'mes': mes,
            'cve_pre': str(cve_pre),
            'nom_umae': str(nom_umae),
            'indicador': indicador_total,
            'valor': float(valor)
        })
    
    wb.close()
    return registros

def insertar_en_sql(registros, conexion_str):
    #Inserción de los regitros con un cursor
    conn = pyodbc.connect(conexion_str)
    cursor = conn.cursor()
    for reg in registros:
        cursor.execute('''
            INSERT INTO eval_UMAE (anio, mes, cve_pre, nom_umae, indicador, valor)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            reg['anio'], reg['mes'], reg['cve_pre'],
            reg['nom_umae'], reg['indicador'], reg['valor']
        ))
    conn.commit()
    conn.close()
    print(f"Se insertaron {len(registros)} registros.")

def procesar_carpeta(carpeta):
    # Obtener lista de archivos .xlsx en la carpeta para insertar todos los reportes dentro de la db
    archivos = [f for f in os.listdir(carpeta) if f.endswith('.xlsx')]
    
    if not archivos:
        print(f"No se encontraron archivos .xlsx en la carpeta {carpeta}")
        return
    
    total_registros = 0
    
    for archivo in archivos:
        ruta_completa = os.path.join(carpeta, archivo)
        print(f"\nProcesando archivo: {archivo}")
        
        try:
            datos = leer_datos(ruta_completa)
            insertar_en_sql(datos, CONEXION_SQL)
            total_registros += len(datos)
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {str(e)}")
            continue
    
    print(f"\nProceso completado. Total de registros insertados: {total_registros}")

if __name__ == "__main__":
    if not os.path.exists(CARPETA_REPORTES):
        print(f"La carpeta {CARPETA_REPORTES} no existe")
    else:
        procesar_carpeta(CARPETA_REPORTES)